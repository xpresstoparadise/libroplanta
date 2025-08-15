#!/usr/bin/env ipy
# -*- coding: UTF-8 -*-

import xml.etree.ElementTree as ET
from openpyxl import Workbook
from nltk import word_tokenize, sent_tokenize
from nltk.corpus import stopwords
from nltk.tokenize.treebank import TreebankWordDetokenizer

def compare_xml(file1, file2):
    # Создаем книгу Excel
    wb = Workbook()
    wb.remove(wb.active)  # Удаляем первый лист
    sheets = {"Удаленные элементы": wb.create_sheet("Удаленные элементы"), "Добавленные элементы": wb.create_sheet("Добавленные элементы"), "Измененные элементы": wb.create_sheet("Измененные элементы")}
    for sheet in sheets.values():
        sheet.append(["Тэг", "Старое значение", "Новое значение", "Атрибут", "Старое значение", "Новое значение"])  # Записываем заголовки столбцов

    tree1 = ET.parse(file1)
    tree2 = ET.parse(file2)

    # Сравниваем тэги и атрибуты
    for el1 in tree1.iter():
        el2 = tree2.find(el1.tag)
        if el2 is None:
            sheets["Удаленные элементы"].append([el1.tag, el1.text, "", "", "", ""])
            for attr, value in el1.attrib.items():
                sheets["Удаленные элементы"].append(["", "", "", attr, value, ""])
        elif el1.text != el2.text:
            sheets["Измененные элементы"].append([el1.tag, el1.text, el2.text, "", "", ""])
            # Сравниваем атрибуты, если элементы совпадают
            attrs = [(attr, el1.get(attr), el2.get(attr)) for attr in set(el1.attrib) | set(el2.attrib)]
            diff_attrs = [a for a in attrs if a[1] != a[2]]
            if diff_attrs:
                for attr, old_value, new_value in diff_attrs:
                    sheets["Измененные элементы"].append(["", "", "", attr, old_value, new_value])

    # Проверяем добавленные и измененные тэги и атрибуты
    for el2 in tree2.iter():
        el1 = tree1.find(el2.tag)
        if el1 is None:
            sheets["Добавленные элементы"].append([el2.tag, "", el2.text, "", "", ""])
            for attr, value in el2.attrib.items():
                sheets["Добавленные элементы"].append(["", "", "", attr, "", value])
        else:
            # Сравниваем атрибуты уже существующих тэгов
            attrs = [(attr, el1.get(attr), el2.get(attr)) for attr in set(el1.attrib) | set(el2.attrib)]
            diff_attrs = [a for a in attrs if a[1] != a[2]]
            if diff_attrs:
                for attr, old_value, new_value in diff_attrs:
                    sheets["Измененные элементы"].append([el2.tag, el1.text, el2.text, attr, old_value, new_value])
                if el2.text != el1.text:
                    sheets["Измененные элементы"].append([el2.tag, el1.text, el2.text, "", "", ""])

    # Сохраняем книгу Excel в файл
    wb.save("comparison_results.xlsx")

def generate_report(file1, file2):
    # Получаем результаты сравнения файлов
    results = compare_xml(file1, file2)

    # Вычисляем количество добавленных, удаленных и измененных тэгов и атрибутов
    num_added_tags = len(results["added_tags"])
    num_added_attrs = sum(len(attrs) for _, attrs in results["added_attrs"].items())
    num_deleted_tags = len(results["deleted_tags"])
    num_deleted_attrs = sum(len(attrs) for _, attrs in results["deleted_attrs"].items())
    num_modified_tags = len(results["modified_tags"])
    num_modified_attrs = sum(len(attrs) for _, attrs in results["modified_attrs"].items())

    # Генерируем список изменений
    modified_changes = []
    for tag, attrs in results["modified_attrs"].items():
        for attr, values in attrs.items():
            old_value, new_value = values
            modified_changes.append(f"- В тэге '{tag}' атрибут '{attr}' был изменен с '{old_value}' на '{new_value}'.")

    # Генерируем текст отчета на основе шаблона
    template = "В результате сравнения файлов {file1} и {file2} были обнаружены следующие изменения:\n" \
               "- Добавлено {num_added_tags} новых тэгов и {num_added_attrs} новых атрибутов.\n" \
               "- Удалено {num_deleted_tags} тэгов и {num_deleted_attrs} атрибутов.\n" \
               "- Изменено {num_modified_tags} тэгов и {num_modified_attrs} атрибутов.\n\n" \
               "Конкретные изменения:\n{modified_changes}"
    report = template.format(file1=file1, file2=file2, num_added_tags=num_added_tags, num_added_attrs=num_added_attrs,
                             num_deleted_tags=num_deleted_tags, num_deleted_attrs=num_deleted_attrs,
                             num_modified_tags=num_modified_attrs, num_modified_attrs=num_modified_attrs, modified_changes='\n'.join(modified_changes))

    # Возвращаем сгенерированный отчет
    return report


file1 = r'C:\Users\xpesstoparadise\Desktop\Новая папка (3)\DMC-21000-A-04-11-00-00A-0B3A-A_001-00_ru-RU.xml'
file2 = r'C:\Users\xpesstoparadise\Desktop\Новая папка (3)\DMC-21000-A-04-11-00-00A-0B3A-A_002-00_ru-RU.xml'

compare_xml(file1, file2)
